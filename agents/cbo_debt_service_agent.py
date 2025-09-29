"""CBO Debt Service Agent.

This module defines an autonomous agent that watches the Congressional Budget Office
(CBO) cost estimate feed, downloads the associated Excel workbooks, computes the
impact of debt service using configurable coefficients, and republishes the
results.  The implementation is designed so that it can run without human
intervention once the required API credentials are supplied via environment
variables.
"""

from __future__ import annotations

import logging
import os
import re
import sqlite3
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
import requests
import tweepy
import yaml
from bs4 import BeautifulSoup
from openpyxl import load_workbook

LOGGER = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Configuration helpers
# ---------------------------------------------------------------------------


def _resolve_path(path: str | Path) -> Path:
    """Resolve a path relative to the project root."""

    expanded = os.path.expandvars(os.path.expanduser(str(path)))
    return Path(expanded).resolve()


@dataclass
class TwitterConfig:
    """Configuration for Twitter access."""

    account_username: str
    follow_screen_name: str
    poll_interval_seconds: int = 300


@dataclass
class FileConfig:
    """Configuration for paths used by the agent."""

    coefficients_csv: Path
    download_dir: Path
    database_path: Path

    @classmethod
    def from_dict(cls, raw: Dict[str, str]) -> "FileConfig":
        return cls(
            coefficients_csv=_resolve_path(raw.get("coefficients_csv", "config/debt_service_coefficients.csv")),
            download_dir=_resolve_path(raw.get("download_dir", "data/cbo_downloads")),
            database_path=_resolve_path(raw.get("database_path", "data/cbo_cost_estimates.db")),
        )


@dataclass
class ExcelConfig:
    """Configuration for editing the Excel workbook."""

    summary_sheet: str = "Summary"
    header_row: int = 1
    total_deficit_label: str = "Total Deficit"
    debt_service_label: str = "Debt service"
    output_suffix: str = "_with_debt_service"


@dataclass
class ProcessingConfig:
    """Configuration for general processing limits."""

    max_tweets_per_run: int = 5


@dataclass
class AgentConfig:
    """Aggregate configuration for the agent."""

    twitter: TwitterConfig
    files: FileConfig
    excel: ExcelConfig = field(default_factory=ExcelConfig)
    processing: ProcessingConfig = field(default_factory=ProcessingConfig)

    @classmethod
    def load(cls, path: str | Path) -> "AgentConfig":
        with open(path, "r", encoding="utf-8") as handle:
            raw = yaml.safe_load(handle) or {}

        twitter_cfg = raw.get("twitter", {})
        file_cfg = raw.get("files", {})
        excel_cfg = raw.get("excel", {})
        processing_cfg = raw.get("processing", {})

        return cls(
            twitter=TwitterConfig(
                account_username=twitter_cfg.get("account_username", "CBO_DebtService"),
                follow_screen_name=twitter_cfg.get("follow_screen_name", "USCBO"),
                poll_interval_seconds=int(twitter_cfg.get("poll_interval_seconds", 300)),
            ),
            files=FileConfig.from_dict(file_cfg),
            excel=ExcelConfig(**excel_cfg),
            processing=ProcessingConfig(**processing_cfg),
        )


# ---------------------------------------------------------------------------
# Data utilities
# ---------------------------------------------------------------------------


class CoefficientTable:
    """Holds the coefficients used to estimate debt service."""

    def __init__(self, coefficients: Dict[str, float], default: float = 0.0):
        self.coefficients = {str(key).strip(): float(value) for key, value in coefficients.items()}
        self.default = float(default)

    @classmethod
    def from_csv(cls, path: Path, default: float = 0.0) -> "CoefficientTable":
        if not path.exists():
            raise FileNotFoundError(f"Coefficient CSV not found: {path}")

        frame = pd.read_csv(path)
        if "label" not in frame.columns or "coefficient" not in frame.columns:
            raise ValueError("Coefficient CSV must have 'label' and 'coefficient' columns")

        coefficients = dict(zip(frame["label"], frame["coefficient"]))
        return cls(coefficients, default)

    def coefficient_for(self, label: str) -> float:
        lookup = str(label).strip()
        return self.coefficients.get(lookup, self.default)


# ---------------------------------------------------------------------------
# Twitter integration
# ---------------------------------------------------------------------------


class TwitterClient:
    """Wrapper around Tweepy to simplify authentication and queries."""

    def __init__(self, config: TwitterConfig):
        self.config = config
        self._client_v2: Optional[tweepy.Client] = None
        self._client_v1: Optional[tweepy.API] = None
        self._follow_user_id: Optional[int] = None

    # Environment variable names for credentials
    BEARER_ENV = "TWITTER_BEARER_TOKEN"
    API_KEY_ENV = "TWITTER_API_KEY"
    API_SECRET_ENV = "TWITTER_API_SECRET"
    ACCESS_TOKEN_ENV = "TWITTER_ACCESS_TOKEN"
    ACCESS_SECRET_ENV = "TWITTER_ACCESS_SECRET"

    def _ensure_clients(self) -> None:
        if self._client_v2 and self._client_v1:
            return

        bearer = os.getenv(self.BEARER_ENV)
        api_key = os.getenv(self.API_KEY_ENV)
        api_secret = os.getenv(self.API_SECRET_ENV)
        access_token = os.getenv(self.ACCESS_TOKEN_ENV)
        access_secret = os.getenv(self.ACCESS_SECRET_ENV)

        missing = [
            env
            for env, value in [
                (self.BEARER_ENV, bearer),
                (self.API_KEY_ENV, api_key),
                (self.API_SECRET_ENV, api_secret),
                (self.ACCESS_TOKEN_ENV, access_token),
                (self.ACCESS_SECRET_ENV, access_secret),
            ]
            if not value
        ]

        if missing:
            raise RuntimeError(
                "Missing Twitter credentials. Set the following environment variables: "
                + ", ".join(missing)
            )

        self._client_v2 = tweepy.Client(
            bearer_token=bearer,
            consumer_key=api_key,
            consumer_secret=api_secret,
            access_token=access_token,
            access_token_secret=access_secret,
            wait_on_rate_limit=True,
        )

        auth = tweepy.OAuth1UserHandler(api_key, api_secret, access_token, access_secret)
        self._client_v1 = tweepy.API(auth)

    @property
    def client_v2(self) -> tweepy.Client:
        self._ensure_clients()
        assert self._client_v2 is not None
        return self._client_v2

    @property
    def client_v1(self) -> tweepy.API:
        self._ensure_clients()
        assert self._client_v1 is not None
        return self._client_v1

    def _resolve_follow_user_id(self) -> int:
        if self._follow_user_id is not None:
            return self._follow_user_id

        user = self.client_v2.get_user(username=self.config.follow_screen_name)
        if not user.data:
            raise RuntimeError(f"Unable to resolve user id for @{self.config.follow_screen_name}")

        self._follow_user_id = user.data.id
        return self._follow_user_id

    def fetch_recent_cost_estimate_tweets(self, limit: int = 10) -> List[tweepy.Tweet]:
        """Fetch recent tweets from the CBO cost estimate account."""

        user_id = self._resolve_follow_user_id()
        response = self.client_v2.get_users_tweets(
            id=user_id,
            max_results=min(limit, 100),
            tweet_fields=["created_at"],
            expansions=["attachments.media_keys", "entities.mentions.username"],
        )

        tweets = list(response.data or [])
        LOGGER.debug("Fetched %d tweets from @%s", len(tweets), self.config.follow_screen_name)
        return tweets

    def follow_primary_account(self) -> None:
        """Ensure that the agent account follows the CBO account."""

        user_id = self._resolve_follow_user_id()
        me = self.client_v2.get_user(username=self.config.account_username)
        if not me.data:
            raise RuntimeError("Unable to resolve agent account")

        self.client_v2.follow(me.data.id, user_id)

    def tweet_with_attachment(self, text: str, attachment_path: Path) -> tweepy.Tweet:
        """Publish a tweet with an uploaded Excel workbook attachment."""

        media = self.client_v1.simple_upload(filename=str(attachment_path))
        response = self.client_v2.create_tweet(text=text, media_ids=[media.media_id_string])
        LOGGER.info("Published tweet %s", response.data.get("id"))
        return response


# ---------------------------------------------------------------------------
# Core agent implementation
# ---------------------------------------------------------------------------


@dataclass
class CostEstimateRecord:
    tweet_id: int
    cbo_url: str
    xlsx_url: str
    local_path: Path
    processed_at: datetime
    debt_service_path: Path


class CBODebtServiceAgent:
    """Agent that automates the full debt service workflow."""

    def __init__(self, config_path: str | Path = "config/cbo_debt_service_agent.yaml") -> None:
        self.config = AgentConfig.load(config_path)
        self.coefficients = CoefficientTable.from_csv(self.config.files.coefficients_csv)
        self.twitter_client = TwitterClient(self.config.twitter)

        self.config.files.download_dir.mkdir(parents=True, exist_ok=True)
        self.config.files.database_path.parent.mkdir(parents=True, exist_ok=True)

        self._initialise_database()

    # ------------------------------------------------------------------
    # Database handling
    # ------------------------------------------------------------------

    def _initialise_database(self) -> None:
        with sqlite3.connect(self.config.files.database_path) as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS cost_estimates (
                    tweet_id INTEGER PRIMARY KEY,
                    cbo_url TEXT NOT NULL,
                    xlsx_url TEXT NOT NULL,
                    local_path TEXT NOT NULL,
                    debt_service_path TEXT NOT NULL,
                    processed_at TEXT NOT NULL
                )
                """
            )
            conn.commit()

    def _record_processed_estimate(self, record: CostEstimateRecord) -> None:
        with sqlite3.connect(self.config.files.database_path) as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO cost_estimates (
                    tweet_id, cbo_url, xlsx_url, local_path, debt_service_path, processed_at
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    record.tweet_id,
                    record.cbo_url,
                    record.xlsx_url,
                    str(record.local_path),
                    str(record.debt_service_path),
                    record.processed_at.isoformat(),
                ),
            )
            conn.commit()

    def _is_tweet_processed(self, tweet_id: int) -> bool:
        with sqlite3.connect(self.config.files.database_path) as conn:
            cursor = conn.execute("SELECT 1 FROM cost_estimates WHERE tweet_id = ?", (tweet_id,))
            return cursor.fetchone() is not None

    # ------------------------------------------------------------------
    # Tweet handling
    # ------------------------------------------------------------------

    URL_REGEX = re.compile(r"https?://\S+")

    def run(self) -> None:
        """Continuously monitor for new tweets and process them."""

        LOGGER.info("Starting CBO Debt Service agent")
        while True:
            try:
                self.process_new_tweets()
            except Exception as exc:  # pragma: no cover - guardrail
                LOGGER.exception("Error while processing tweets: %s", exc)

            LOGGER.debug("Sleeping for %s seconds", self.config.twitter.poll_interval_seconds)
            time.sleep(self.config.twitter.poll_interval_seconds)

    def process_new_tweets(self) -> List[CostEstimateRecord]:
        """Process recently published cost estimate tweets."""

        tweets = self.twitter_client.fetch_recent_cost_estimate_tweets(
            limit=self.config.processing.max_tweets_per_run
        )
        records: List[CostEstimateRecord] = []

        for tweet in tweets:
            if self._is_tweet_processed(tweet.id):
                LOGGER.debug("Skipping tweet %s (already processed)", tweet.id)
                continue

            url = self._extract_first_url(tweet.text or "")
            if not url:
                LOGGER.warning("Tweet %s does not contain a URL", tweet.id)
                continue

            LOGGER.info("Processing tweet %s", tweet.id)
            try:
                record = self._process_tweet(tweet.id, url)
                records.append(record)
            except Exception:
                LOGGER.exception("Failed to process tweet %s", tweet.id)

        return records

    def _extract_first_url(self, text: str) -> Optional[str]:
        match = self.URL_REGEX.search(text)
        return match.group(0) if match else None

    # ------------------------------------------------------------------
    # Cost estimate processing
    # ------------------------------------------------------------------

    def _process_tweet(self, tweet_id: int, url: str) -> CostEstimateRecord:
        cbo_url, xlsx_url = self._resolve_cost_estimate_urls(url)
        local_path = self._download_xlsx(xlsx_url)
        debt_service_path = self._augment_workbook_with_debt_service(local_path)

        text = f"Debt Service for Cost Estimate {Path(local_path).stem}"
        self.twitter_client.tweet_with_attachment(text=text, attachment_path=debt_service_path)

        record = CostEstimateRecord(
            tweet_id=tweet_id,
            cbo_url=cbo_url,
            xlsx_url=xlsx_url,
            local_path=local_path,
            debt_service_path=debt_service_path,
            processed_at=datetime.utcnow(),
        )
        self._record_processed_estimate(record)
        return record

    def _resolve_cost_estimate_urls(self, tweet_url: str) -> Tuple[str, str]:
        """Follow the tweet URL to discover the XLSX download link."""

        LOGGER.debug("Fetching cost estimate page: %s", tweet_url)
        response = requests.get(tweet_url, timeout=60)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

        # Attempt to identify the canonical CBO cost estimate URL
        canonical = soup.find("link", rel="canonical")
        cbo_url = canonical["href"] if canonical and canonical.has_attr("href") else tweet_url

        # Find the first Excel link on the page
        xlsx_link = soup.find("a", href=re.compile(r"\.xlsx$", re.IGNORECASE))
        if not xlsx_link:
            raise RuntimeError("Unable to locate XLSX download link on CBO page")

        xlsx_url = requests.compat.urljoin(cbo_url, xlsx_link["href"])  # type: ignore[index]
        return cbo_url, xlsx_url

    def _download_xlsx(self, url: str) -> Path:
        LOGGER.debug("Downloading XLSX from %s", url)
        response = requests.get(url, timeout=120)
        response.raise_for_status()

        filename = url.split("/")[-1]
        target = self.config.files.download_dir / filename
        with open(target, "wb") as handle:
            handle.write(response.content)

        LOGGER.info("Saved cost estimate to %s", target)
        return target

    def _augment_workbook_with_debt_service(self, path: Path) -> Path:
        wb = load_workbook(path)
        sheet_name = self.config.excel.summary_sheet

        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Workbook does not contain expected sheet '{sheet_name}'")

        sheet = wb[sheet_name]
        header_row_index = self.config.excel.header_row
        headers = [cell.value for cell in sheet[header_row_index]]

        # The first column is expected to contain row labels; remaining columns are per-year values.
        data_headers = []
        for header in headers[1:]:
            if header is None:
                data_headers.append("")
            else:
                data_headers.append(str(header))

        deficit_row_index, deficit_values = self._locate_total_deficit_row(sheet)
        data_headers = data_headers[: len(deficit_values)]
        debt_service_values = self._calculate_debt_service_values(data_headers, deficit_values)

        insert_at = deficit_row_index + 1
        sheet.insert_rows(insert_at)
        sheet.cell(row=insert_at, column=1, value=self.config.excel.debt_service_label)
        for offset, value in enumerate(debt_service_values, start=2):
            sheet.cell(row=insert_at, column=offset, value=value)

        output_path = path.with_name(f"{path.stem}{self.config.excel.output_suffix}{path.suffix}")
        wb.save(output_path)
        LOGGER.info("Augmented workbook saved to %s", output_path)
        return output_path

    def _locate_total_deficit_row(self, sheet) -> Tuple[int, List[float]]:
        label = self.config.excel.total_deficit_label.lower()
        header_row_index = self.config.excel.header_row
        first_data_row = header_row_index + 1

        for row_idx in range(first_data_row, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value and label in str(cell_value).lower():
                row_values = []
                for col_idx in range(2, sheet.max_column + 1):
                    value = sheet.cell(row=row_idx, column=col_idx).value
                    row_values.append(value)
                return row_idx, row_values

        raise RuntimeError(f"Could not find a row containing '{self.config.excel.total_deficit_label}'")

    def _calculate_debt_service_values(self, headers: Sequence[str], deficit_values: Sequence) -> List[Optional[float]]:
        results: List[Optional[float]] = []
        for header, value in zip(headers, deficit_values):
            if value in (None, ""):
                results.append(None)
                continue

            try:
                numeric_value = float(value)
            except (TypeError, ValueError):
                results.append(None)
                continue

            coefficient = self.coefficients.coefficient_for(header)
            results.append(round(numeric_value * coefficient, 2))

        return results


# ---------------------------------------------------------------------------
# Command line entry point
# ---------------------------------------------------------------------------


def main() -> None:  # pragma: no cover - integration entry point
    logging.basicConfig(level=logging.INFO)
    agent = CBODebtServiceAgent()
    agent.run()


if __name__ == "__main__":  # pragma: no cover
    main()
